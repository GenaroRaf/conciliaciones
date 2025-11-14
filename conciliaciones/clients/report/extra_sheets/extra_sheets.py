from io import BytesIO

from k_link.db.core import ObjectId
from k_link.db.daos import ERPFilesDAO, LinkServicesDAO, ProjectDAO
from k_link.db.models import LinkServices
from k_link.extensions.conciliation_type import ConciliationType
from k_link.extensions.report_config import ReportConfig
from loggerk import LoggerK
from openpyxl import Workbook, load_workbook

from conciliaciones.clients.report.styles.report_styles import ReportStyles
from conciliaciones.clients.report.utils.report_data_handler import ReportDataHandler
from conciliaciones.clients.report.utils.report_sheets import ReportSheets
from conciliaciones.utils.completion_handler.airflow_contex_exception import (
    AirflowContexException,
)
from conciliaciones.utils.data.normalize import normalize_date
from conciliaciones.utils.redis.redis_keys import RedisKeys
from conciliaciones.utils.redis.redis_storage import RedisStorage


class ExtraSheets:
    """
    * Clase encargada de crear hojas adicionales en el archivo de Excel.
    * Integra datos de reportes extras configurados y se les aplica formato de SAT
    """

    def __init__(
        self,
        run_id: str,
        project_id_str: str,
        month: int,
        year: int,
        conciliation_type: ConciliationType,
    ) -> None:
        """
        * Inicializa la clase con los datos necesarios para generar las hojas extras.
        TODO @param run_id: ID de ejecución del reporte.
        TODO @param project_id_str: ID del proyecto en formato string.
        TODO @param month: Mes del periodo conciliado.
        TODO @param year: Año del periodo conciliado.
        TODO @param conciliation_type: Tipo de conciliación.
        TODO @param tipo_reporte: Tipo de reporte a generar.
        """

        self._logger = LoggerK(self.__class__.__name__)

        self._project_dao = ProjectDAO()
        self._erp_files_dao = ERPFilesDAO()
        self._link_services_dao = LinkServicesDAO()

        self.project_id = ObjectId(project_id_str)
        self.redis = RedisStorage()
        self.month = month
        self.year = year
        self._conciliation_type: ConciliationType = conciliation_type

        self._reporte_data = ReportDataHandler(
            run_id=run_id,
            project_id_str=project_id_str,
            month=month,
            year=year,
            conciliation_type=conciliation_type,
        )
        self._redis_key = RedisKeys(
            run_id=run_id,
            project_id_str=project_id_str,
            month=month,
            year=year,
            conciliation_type=conciliation_type,
        )
        self._airflow_fail_exception = AirflowContexException(
            year=year,
            month=month,
            project_id=project_id_str,
            run_id=run_id,
            conciliation_type=conciliation_type,
        )

        self._report_sheets = ReportSheets(
            run_id=run_id,
            project_id_str=project_id_str,
            month=month,
            year=year,
            conciliation_type=conciliation_type,
        )
        self._report_styles = ReportStyles(
            run_id=run_id,
            project_id_str=project_id_str,
            month=month,
            year=year,
            conciliation_type=conciliation_type,
        )

    async def create_extra_sheets(self) -> None:
        """
        * Crea las hojas adicionales en el archivo de Excel almacenado en Redis.
        * Integra los datos de reportes extras configurados y les aplica formato de SAT.
        ! Si el Excel buffer no existe en Redis, se lanza una excepción.
        """
        link_services: LinkServices | None = await self._link_services_dao.get(
            project_id=ObjectId(self.project_id)
        )

        if link_services is None:
            self._airflow_fail_exception.handle_and_store_exception(
                f"No se encontro configuración de LinkServices para el proyecto: {self.project_id}"
            )

        report_config: ReportConfig | None = link_services.report_config

        if report_config is None or not report_config.report_sheets:
            self._logger.error(
                f"Report config o extra sheets vacias, proyecto: {self.project_id}"
            )
            return

        today = normalize_date(self.year, self.month)
        self._logger.info(f"ProjectId: {self.project_id}   Fecha: {today} ")

        excel_buffer_key: str = self._redis_key.get_excel_buffer_key()
        excel_buffer: BytesIO | None = self.redis.get(
            key=excel_buffer_key, object_type=BytesIO
        )

        if excel_buffer is None:
            self._airflow_fail_exception.handle_and_store_exception(
                f"No se encontró el buffer de Excel para el proyecto: {self.project_id} en la hoja Extra Sheets."
            )

        excel_buffer.seek(0)
        workbook: Workbook = load_workbook(excel_buffer)

        # Extra Sheet SAT sin ERP
        (
            df_sheets,
            sheet_names,
            report_sheets_k,
        ) = await self._report_sheets.get_report_sheets()

        for sheet, name, report_mongo in zip(df_sheets, sheet_names, report_sheets_k):
            if sheet.empty:
                self._logger.info(f"Dataframe vacio, no se crea la hoja {name}")
                continue

            # Verifica si la hoja ya existe y la elimina si es necesario
            if name in workbook.sheetnames:
                del workbook[name]

            worksheet = workbook.create_sheet(title=name)

            self._logger.info(f"Procesando hoja: {name}")

            self._logger.info(f"DataFrame {name}: {sheet.columns}")

            self._logger.info(f"Formato de hoja: {name}")

            # Escribir los encabezados y datos manualmente
            for r_idx, row in enumerate(sheet.to_records(index=False), start=2):
                for c_idx, value in enumerate(row, start=1):
                    # Convertir listas, dicts o sets a string legible
                    if isinstance(value, list):
                        value = ", ".join(
                            (
                                "; ".join(f"{k}: {v}" for k, v in item.items())
                                if isinstance(item, dict)
                                else str(item)
                            )
                            for item in value
                        )
                    elif isinstance(value, dict):
                        value = "; ".join(f"{k}: {v}" for k, v in value.items())
                    elif isinstance(value, set):
                        value = ", ".join(str(v) for v in value)

                    worksheet.cell(row=r_idx, column=c_idx, value=value)  # type: ignore

            await self._report_styles.headers_extra_sheets(
                worksheet=worksheet, df=sheet, report_id=report_mongo.report_id
            )

            self._logger.info(f"Extra Sheet: {name} {sheet}")

        # Se guarda el excel buffer
        self._logger.info(
            f"Guardando buffer de Excel de las ExtraSheets en Redis: {excel_buffer}"
        )

        new_excel_buffer = BytesIO()
        workbook.save(new_excel_buffer)
        new_excel_buffer.seek(0)

        # Guarda el nuevo buffer en Redis
        self.redis.set(key=excel_buffer_key, value=new_excel_buffer)
