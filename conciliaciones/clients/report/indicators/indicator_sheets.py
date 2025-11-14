# pylint: disable=too-many-instance-attributes
# reason: This class requires multiple instance attributes to store
# configuration and state information as part of the data model.
# Reducing the number would harm clarity or cohesion of the design.

from io import BytesIO

import pandas as pd
from k_link.db.core import ObjectId
from k_link.db.daos import IndicatorCustomDAO, IndicatorTemplateDAO
from k_link.db.models import IndicatorCustom, IndicatorTemplate
from k_link.extensions.conciliation_type import ConciliationType
from k_link.extensions.indicators import (
    IndicatorConfig,
    IndicatorType,
    OperationIndicator,
)
from k_link.extensions.report_config import TipoDato
from loggerk import LoggerK
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from conciliaciones.clients.report.indicators.indicator_handler import IndicatorHandler
from conciliaciones.utils.completion_handler.airflow_contex_exception import (
    AirflowContexException,
)
from conciliaciones.utils.redis.redis_keys import RedisKeys
from conciliaciones.utils.redis.redis_storage import RedisStorage


class Indicator:
    def __init__(
        self,
        run_id: str,
        project_id_str: str,
        year: int,
        month: int,
        conciliation_type: ConciliationType,
    ) -> None:
        self._logger = LoggerK(self.__class__.__name__)
        self._run_id: str = run_id

        self._project_id_str: str = project_id_str
        self._year: int = year
        self._month: int = month
        self._conciliation_type: ConciliationType = conciliation_type

        self._indicator_templates_dao = IndicatorCustomDAO()
        self._indicador_dao = IndicatorTemplateDAO()
        self._redis = RedisStorage()
        self._redis_keys = RedisKeys(
            run_id=self._run_id,
            project_id_str=self._project_id_str,
            month=self._month,
            year=self._year,
            conciliation_type=self._conciliation_type,
        )
        self._airflow_fail_exception = AirflowContexException(
            year=year,
            month=month,
            project_id=project_id_str,
            run_id=run_id,
            conciliation_type=conciliation_type,
        )

        # Estilos para numeros
        self.style_number = "#,##0"
        self.style_currency = "$#,##0.00"

        # Colores de letra
        self.style_black = Font(color="000000")  # Negro
        self.style_blue = Font(color="0000FF")  # Azul
        self.style_red = Font(color="FF0000")  # Rojo

    # Estilos para las celdas
    bold_font = Font(bold=True)
    black_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )

    async def get_configuration_indicators(self) -> None:
        indicator_templates: (
            IndicatorCustom | None
        ) = await self._indicator_templates_dao.get(
            project_id=ObjectId(self._project_id_str)
        )

        if indicator_templates is None:
            self._airflow_fail_exception.handle_and_store_exception(
                f"No se encontraron plantillas de indicadores para el proyecto: {self._project_id_str}"
            )

        indicators: list[IndicatorConfig] = indicator_templates.indicadores

        if not indicators:
            return

        df_erp_sat: pd.DataFrame | None = self._redis.get_df(
            redis_key=self._redis_keys.get_sat_erp_redis_key()
        )

        if df_erp_sat is None:
            self._airflow_fail_exception.handle_and_store_exception(
                f"No se encontró DataFrame para la llave de Redis: {self._redis_keys.get_sat_erp_redis_key()}"
            )

        excel_buffer: BytesIO | None = self._redis.get(
            key=self._redis_keys.get_excel_buffer_key(), object_type=BytesIO
        )

        if excel_buffer is None:
            self._airflow_fail_exception.handle_and_store_exception(
                f"No se encontró el buffer de Excel para el proyecto: {self._project_id_str} en la hoja de Indicadores."
            )

        excel_buffer.seek(0)
        workbook: Workbook = load_workbook(excel_buffer)

        for indicator_config in indicators:
            indicator_collection: (
                IndicatorTemplate | None
            ) = await self._indicador_dao.get_by_id(
                item_id=indicator_config.indicator_id
            )

            if indicator_collection is None:
                self._airflow_fail_exception.handle_and_store_exception(
                    f"No se encontró el indicador con ID: {indicator_config.indicator_id}"
                )

            indicador_handler: IndicatorHandler = IndicatorHandler(
                df=df_erp_sat,
                indicator_config=indicator_config,
                project_id_str=self._project_id_str,
                year=self._year,
                month=self._month,
                conciliation_type=self._conciliation_type,
                run_id=self._run_id,
            )

            df_indicador, sheet_name = indicador_handler.apply_indicator()

            if df_indicador is None or df_indicador.empty:
                continue

            if sheet_name in workbook.sheetnames:
                del workbook[sheet_name]

            sheet = workbook.create_sheet(title=sheet_name)

            for r_idx, row in enumerate(
                dataframe_to_rows(df_indicador, index=False, header=True), start=1
            ):
                for c_idx, value in enumerate(row, start=1):
                    sheet.cell(row=r_idx, column=c_idx, value=value)  # type: ignore

            # Aplica estilos
            await self.apply_style_indicators(
                sheet=sheet,
                df_indicador=df_indicador,
                sheet_name=indicator_collection.sheet_name,
                indicator=indicator_collection,
            )

            self._logger.info(
                f"Indicador: {indicator_collection.sheet_name} {df_indicador}"
            )

            indicator_name = indicator_collection.sheet_name

            redis_key = self._redis_keys.get_indicators_redis_key(
                indicator_name=indicator_name
            )

            buffer_df: BytesIO = self._redis.set_parquet(df=df_indicador)
            self._redis.set(
                key=redis_key,
                value=buffer_df,
            )
            self._redis.medir_tamano_valor(df=df_indicador, buffer_df=buffer_df)

            self._logger.info(f"Indicador guardado en Redis: {redis_key}")

        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)

        self._redis.set(
            key=self._redis_keys.get_excel_buffer_key(),
            value=excel_buffer,
        )

    async def apply_style_indicators(
        self,
        sheet,
        df_indicador: pd.DataFrame,
        sheet_name: str,
        indicator: IndicatorTemplate,
    ) -> None:
        """
        Aplica estilo a la tabla de indicadores.

        Args:
            sheet: Hoja de Excel donde aplicar el estilo
            df_inficador: DataFrame con los datos del indicador
            sheet_name: Nombre de la hoja
        """
        self._logger.info(f"Aplicando estilo a la hoja: {sheet_name}")
        style_number_map = {
            TipoDato.NUMERO.value: self.style_number,
            TipoDato.MONEDA.value: self.style_currency,
        }

        style_letter_map = {
            0: self.style_black,
            1: self.style_blue,
            -1: self.style_red,
        }

        # Obtener dimensiones del DataFrame
        num_rows = len(df_indicador)
        num_cols = len(df_indicador.columns)

        # Estilo para headers (azul como en FormatoPivotTable)
        header_style = "0075AF"  # Azul
        header_font_color = "FFFFFF"

        # Aplicar estilo a los headers (primera fila)
        for col in range(1, num_cols + 1):
            cell = sheet.cell(row=1, column=col)
            self._apply_cell_style_header(cell, header_style, header_font_color)

        operation_indicators: list[list[OperationIndicator]] = [
            operation_column.operation_indicators
            for operation_column in indicator.operation_columns
        ]
        row_styles: list[list[str]] = []
        row_styles_letter: list[list[int]] = []

        for list_operations in operation_indicators:
            style_column: list[str] = []
            style_letter: list[int] = []
            for operation in list_operations:
                style_column.append(operation.data_type)
                style_letter.append(operation.mult_operation.vertical)
            row_styles.append(style_column)
            row_styles_letter.append(style_letter)

        # Estilos para la columna total
        if indicator.indicator_type.value != IndicatorType.RESUMEN.value:
            style_column = []
            style_letter = []
            for _ in range(len(row_styles[0])):
                style_column.append("MONEDA")

            for _ in range(len(row_styles_letter[0])):
                style_letter.append(0)

            row_styles.append(style_column)
            row_styles_letter.append(style_letter)

            for col, style_column in enumerate(row_styles):
                style_column.append("MONEDA")

            for col, style_letter in enumerate(row_styles_letter):
                style_letter.append(0)

        cont = 0
        if indicator.indicator_type.value != IndicatorType.RESUMEN.value:
            for col in range(1, num_cols + 1):
                cell = sheet.cell(row=num_rows + 1, column=col)
                cell.font = self.bold_font

            for row in range(2, num_rows + 2):
                cell = sheet.cell(row=row, column=num_cols)
                cell.font = self.bold_font

        for row in range(2, num_rows + 2):
            for col in range(1, num_cols + 1):
                cell = sheet.cell(row=row, column=col)

                cell.border = self.black_border
            cont += 1

        cont = 0
        cont_styles = 0
        for row in range((len(indicator.custom_indicators) + 1), num_cols + 1):
            for col in range(2, num_rows + 2):
                cell_values = sheet.cell(row=col, column=row)
                style_numbers: list[str] = row_styles[cont]
                style_value_number = style_numbers[cont_styles]

                style_letters: list[int] = row_styles_letter[cont]
                style_value_letter = style_letters[cont_styles]

                if style_value_number in style_number_map:
                    cell_values.number_format = style_number_map[style_value_number]

                if style_value_letter in style_letter_map:
                    cell_values.font = style_letter_map[style_value_letter]
                cont_styles += 1
            cont += 1
            cont_styles = 0

        self._logger.info(f"Estilo aplicado exitosamente a la hoja: {sheet_name}")

    def _apply_cell_style_header(self, cell, bg_color: str, font_color: str) -> None:
        """
        Aplica estilo a una celda específica.

        Args:
            cell: Celda de Excel
            bg_color: Color de fondo en hexadecimal
            font_color: Color de fuente en hexadecimal
        """
        cell.fill = PatternFill(
            start_color=bg_color, end_color=bg_color, fill_type="solid"
        )
        cell.font = Font(color=font_color, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Side(border_style="thin", color="000000")
        cell.border = Border(
            left=thin_border, right=thin_border, top=thin_border, bottom=thin_border
        )

    async def get_df_redis(self) -> pd.DataFrame:
        redis_key = self._redis_keys.get_sat_erp_redis_key()

        df_base: pd.DataFrame | None = self._redis.get_df(redis_key=redis_key)

        if df_base is None:
            self._airflow_fail_exception.handle_and_store_exception(
                f"No se encontró DataFrame para la llave de Redis: {redis_key}"
            )

        return df_base
