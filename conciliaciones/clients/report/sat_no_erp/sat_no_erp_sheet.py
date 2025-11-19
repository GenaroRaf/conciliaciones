from io import BytesIO

from k_link.db.core import ObjectId
from k_link.db.daos import ERPFilesDAO, ProjectDAO
from k_link.db.models.project import Project
from k_link.extensions.conciliation_type import ConciliationType
from loggerk import LoggerK
from openpyxl import Workbook, load_workbook

from conciliaciones.clients.report.styles.report_styles import ReportStyles
from conciliaciones.clients.report.utils.report_data_handler import ReportDataHandler
from conciliaciones.utils.data.normalize import normalize_date
from conciliaciones.utils.redis.redis_keys import RedisKeys
from conciliaciones.utils.redis.redis_storage import RedisStorage


class SATNoERPSheet:
    """
    * Clase encargada de crear la hoja SAT no ERP en el archivo de Excel.
    * Integra los datos de facturas del SAT que no están en ERP y les aplica
    """

    _sheet_name: str = "SAT no ERP"

    def __init__(
        self,
        run_id: str,
        project_id_str: str,
        month: int,
        year: int,
        conciliation_type: ConciliationType,
        tipo_reporte: str | None,
    ) -> None:
        """
        * Inicializa la clase con los datos necesarios para generar la hoja SAT no ERP.

        TODO @param run_id: ID de ejecución del reporte.
        TODO @param project_id_str: ID del proyecto en formato string.
        TODO @param month: Mes del periodo conciliado.
        TODO @param year: Año del periodo conciliado.
        TODO @param conciliation_type: Tipo de conciliación.
        TODO @param tipo_reporte: Tipo de reporte a generar.
        """
        self._logger = LoggerK(self.__class__.__name__)

        self._project_dao = ProjectDAO()
        self._erp_files_dao = ERPFilesDAO()

        self.project_id = ObjectId(project_id_str)
        self.redis = RedisStorage()
        self.month = month
        self.year = year
        self._conciliation_type: ConciliationType = conciliation_type
        self._tipo_reporte = tipo_reporte

        self._reporte_data = ReportDataHandler(
            run_id=run_id,
            project_id_str=project_id_str,
            month=month,
            year=year,
            conciliation_type=conciliation_type,
        )
        self._redis_key = RedisKeys(
            run_id=run_id,
            project_id_str=project_id_str,
            month=month,
            year=year,
            conciliation_type=conciliation_type,
        )
        self._report_styles = ReportStyles(
            run_id=run_id,
            project_id_str=project_id_str,
            month=month,
            year=year,
            conciliation_type=conciliation_type,
        )

        if tipo_reporte is None:
            project: Project | None = self._project_dao.get_by_id_sync(
                item_id=self.project_id
            )

            if project is None:
                raise ValueError(f"No existe proyecto con id: {self.project_id}")

            self.tipo_reporte = project.project_type
        else:
            self.tipo_reporte = tipo_reporte

        self._logger.info(
            f"Tipo de proyecto: {self.tipo_reporte} para tipo {self._conciliation_type.value}"
        )

        self._is_suppliers: bool = False

        if self.tipo_reporte == "Proveedores":
            self._is_suppliers = True

    async def create_sat_no_erp_sheet(self) -> None:
        """
        * Crea una hoja de Excel con los datos de facturas del SAT que no están en ERP.
        TODO @param is_suppliers: Indica si el reporte es de proveedores (en cuyo caso no se genera esta hoja).
        ! Si el Excel buffer no existe en Redis, se lanza una excepción.
        """

        if (
            not self._is_suppliers
            or self._conciliation_type == ConciliationType.UNITARY
        ):
            self._logger.info(
                f"No se generará la hoja {self._sheet_name} porque no es un proveedor o es una conciliación unitaria."
            )
            return

        today = normalize_date(self.year, self.month)
        self._logger.info(f"ProjectId: {self.project_id}   Fecha: {today} ")

        excel_buffer_key: str = self._redis_key.get_excel_buffer_key()
        excel_buffer: BytesIO | None = self.redis.get(
            key=excel_buffer_key, object_type=BytesIO
        )

        if excel_buffer is None:
            raise RuntimeError("Excel buffer not found in Redis.")

        excel_buffer.seek(0)
        workbook: Workbook = load_workbook(excel_buffer)

        # Verifica si la hoja ya existe y la elimina si es necesario
        if self._sheet_name in workbook.sheetnames:
            del workbook[self._sheet_name]

        worksheet = workbook.create_sheet(title=self._sheet_name)

        (
            df_sat_no_erp,
            headers_sat_config,
        ) = await self._reporte_data.get_sat_no_erp_data()

        self._logger.info(f"df Sat despues del filtro: {df_sat_no_erp}")

        if df_sat_no_erp.empty or df_sat_no_erp is None:
            self._logger.info(f"No hay datos para la hoja {self._sheet_name}")
            return

        self._logger.info(f"DataFrame SAT no ERP: {df_sat_no_erp.columns}")
        self._logger.info(f"Formato de hoja: {self._sheet_name}")

        # Escribir los encabezados y datos manualmente
        for r_idx, row in enumerate(df_sat_no_erp.to_records(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                # Convertir listas, dicts o sets a string legible
                if isinstance(value, list):
                    value = ", ".join(
                        (
                            "; ".join(f"{k}: {v}" for k, v in item.items())
                            if isinstance(item, dict)
                            else str(item)
                        )
                        for item in value
                    )
                elif isinstance(value, dict):
                    value = "; ".join(f"{k}: {v}" for k, v in value.items())
                elif isinstance(value, set):
                    value = ", ".join(str(v) for v in value)

                worksheet.cell(row=r_idx, column=c_idx, value=value)

        await self._report_styles.headers_sat(
            worksheet=worksheet,
            df=df_sat_no_erp,
            erp_sat=False,
            headers_sat_config=headers_sat_config,
        )

        # Se guarda el excel buffer
        self._logger.info(
            f"Guardando buffer de Excel del {self._sheet_name} en Redis: {excel_buffer}"
        )

        self._logger.info(f"SAT no ERP: {df_sat_no_erp}")

        new_excel_buffer = BytesIO()
        workbook.save(new_excel_buffer)
        new_excel_buffer.seek(0)

        # Guarda el nuevo buffer en Redis
        self.redis.set(key=excel_buffer_key, value=new_excel_buffer)
