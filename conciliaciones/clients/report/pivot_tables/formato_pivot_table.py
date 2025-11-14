from collections import Counter
from io import BytesIO

from k_link.db.core import ObjectId
from k_link.db.daos.pivot_templates_dao import PivotCustomDAO
from k_link.db.models import PivotTable
from k_link.db.models.pivot_customs import PivotCustom
from k_link.extensions.conciliation_type import ConciliationType
from k_link.extensions.pipeline.pivot_table_options import PivotOptions
from loggerk import LoggerK
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pandas import DataFrame

from conciliaciones.clients.report.utils.report_data_handler import ReportDataHandler
from conciliaciones.utils.completion_handler.airflow_contex_exception import (
    AirflowContexException,
)
from conciliaciones.utils.redis.redis_keys import RedisKeys
from conciliaciones.utils.redis.redis_storage import RedisStorage


class FormatoPivotTable:
    reporte_data: ReportDataHandler

    def __init__(
        self,
        run_id: str,
        project_id_str: str,
        month: int,
        year: int,
        conciliation_type: ConciliationType,
    ) -> None:
        """
        Inicializa la clase ReportManager.

        Args:
            _logger(LoggerK): instancia de clase logger
            project_id_str (str): ID del proyecto.
            project_id(ObjectId): ID del proyecto
            redis: instancia redis
            month (str): Mes del reporte.
            year (str): Año del reporte.
            redis_keys(RedisKeys): llaves de acceso a redis.
        """

        self._logger = LoggerK(self.__class__.__name__)
        self.run_id: str = run_id
        self.project_id_str = project_id_str
        self.project_id = ObjectId(project_id_str)
        self.redis = RedisStorage()
        self.month: int = month
        self.year: int = year
        self._conciliation_type = conciliation_type

        self._pivot_template_dao = PivotCustomDAO()

        self.redis_keys = RedisKeys(
            run_id=run_id,
            project_id_str=self.project_id_str,
            month=self.month,
            year=self.year,
            conciliation_type=self._conciliation_type,
        )
        self._airflow_fail_exception = AirflowContexException(
            year=year,
            month=month,
            project_id=project_id_str,
            run_id=run_id,
            conciliation_type=conciliation_type,
        )

    async def aplicar_formato(self) -> None:
        """Obtiene la lista aggfunc correspondinete de pivot_template"""
        pivot_customs: PivotCustom | None = self._pivot_template_dao.get_sync(
            project_id=self.project_id
        )

        if not pivot_customs:
            self._logger.info(
                f"Pivot custom para proyecto: {self.project_id} no existe."
            )
            return

        if pivot_customs.pivot_tables == []:
            self._logger.info(
                f"No hay tablas dinámicas para el proyecto: {self.project_id}."
            )
            return

        # Instancia de report_handler
        self.get_report_data_handler()

        # Recupero formato de configuracion de la pivot table
        format_config_key: str = self.redis_keys.get_pivot_table_format_config_key()
        format_config_pivot_table: dict[str, tuple[DataFrame, PivotTable]] | None = (
            self.redis.get(
                key=format_config_key,
                object_type=dict[str, tuple[DataFrame, PivotTable]],
            )
        )

        # Recupero excel buffer
        excel_buffer_key: str = self.redis_keys.get_excel_buffer_key()
        excel_buffer: BytesIO | None = self.redis.get(
            key=excel_buffer_key, object_type=BytesIO
        )

        if excel_buffer is None:
            self._airflow_fail_exception.handle_and_store_exception(
                f"No se encontró el buffer de Excel para el proyecto: {self.project_id} para el formato de Tablas Dinamicas."
            )

        excel_buffer.seek(0)
        workbook: Workbook = load_workbook(excel_buffer)

        if not format_config_pivot_table:
            self._logger.error(
                "No se generaron tablas dinámicas. Solo se aplicará formato a las hojas existentes."
            )
        else:
            self._logger.info("Aplicando formato a tablas dinámicas.")
            for name, (df_pt, pivot_table_config) in format_config_pivot_table.items():
                self._logger.info(f"reset index: {pivot_table_config.reset_index}")
                await self.format_pivot_table_sheet(
                    workbook=workbook,
                    sheet_name=name,
                    df_pivot=df_pt,
                    pivot_table_config=pivot_table_config,
                )

        new_excel_buffer = BytesIO()
        workbook.save(new_excel_buffer)
        new_excel_buffer.seek(0)

        # Guarda el nuevo buffer en Redis
        self.redis.set(key=excel_buffer_key, value=new_excel_buffer)

    def get_report_data_handler(self) -> None:
        """Obtiene el manejador de datos del reporte."""
        self.reporte_data = ReportDataHandler(
            run_id=self.run_id,
            project_id_str=self.project_id_str,
            month=self.month,
            year=self.year,
            conciliation_type=self._conciliation_type,
        )

    bold_font = Font(bold=True)
    black_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )

    async def format_pivot_table_sheet(
        self, workbook, sheet_name, df_pivot, pivot_table_config
    ) -> None:
        """Formato a pivot table."""
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"La hoja '{sheet_name}' no existe en el archivo.")

        sheet = workbook[sheet_name]
        rangos_combinados = sheet.merged_cells.ranges

        self._logger.info(f"Hoja: {sheet_name}")
        style: str = ""

        filter_pivot_custom: PivotOptions | None = self._get_aggfunc(pivot_table_config)

        # Formato de los valores
        filas_columns: int = len(df_pivot.columns.names)
        filas_index: int = len(df_pivot.index.names)
        ultimos: int = len(df_pivot.columns)
        combinada: bool = False
        rangos_validos = []

        if filter_pivot_custom is None:
            num_columns = len(df_pivot.columns)
            num_index = len(df_pivot.index)
        else:
            num_columns = len(filter_pivot_custom.columns_params)
            num_index = len(filter_pivot_custom.index_params)

        # Celdas mergedas
        for rango in rangos_combinados:
            for row in sheet.iter_rows(
                min_row=rango.min_row,
                max_row=rango.max_row,
                min_col=rango.min_col,
                max_col=rango.max_col,
            ):
                for cell in row:
                    if cell.row <= filas_columns:
                        combinada = True

            if combinada:
                rangos_validos.append(rango)

            combinada = False

        if rangos_validos and num_columns != 0:
            for rango in rangos_validos:
                for row in sheet.iter_rows(
                    min_row=rango.min_row,
                    max_row=rango.max_row,
                    min_col=rango.min_col,
                    max_col=rango.max_col,
                ):
                    for cell in row:
                        if cell.row <= filas_columns:
                            if cell.row == 1:
                                style = "43C9AF"
                                font_color = "000000"
                            else:
                                style = "0075AF"
                                font_color = "FFFFFF"

                            self.apply_style(cell, style=style, font_color=font_color)
        elif num_columns != 0:
            style = "43C9AF"
            font_color = "000000"

            for columna in range(filas_index + 1, (filas_index + 1) + ultimos):
                cell = sheet.cell(row=1, column=columna)

                self.apply_style(cell, style=style, font_color=font_color)

        style = "0075AF"
        col_num = 0
        start_fil: int = filas_columns

        if filas_columns > 1:
            col_num = 1
            start_fil = 2
        else:
            start_fil = 1
        font_color = "FFFFFF"

        # Formato filas (columna 3) cuando no hay reset index
        if not pivot_table_config.reset_index:
            for column in range(1, filas_index + 1):
                cell = sheet.cell(row=filas_columns + col_num, column=column)

                self.apply_style(cell, style=style, font_color=font_color)

        self._logger.info(f"Ultimos: {ultimos}")

        if pivot_table_config.reset_index:
            columnas_multindex = num_index + num_columns
        else:
            columnas_multindex = filas_index

        # Formato tamaño multindex (valores de las columnas)
        for i in range(start_fil, filas_columns + 1):
            for column in range(columnas_multindex + 1, filas_index + 1 + ultimos):
                cell = sheet.cell(row=i, column=column)

                self.apply_style(cell, style=style, font_color=font_color)

        # Formato columnas
        if not pivot_table_config.reset_index:  # Sin reset index
            for fila in range(2, filas_columns + 1):
                cell = sheet.cell(row=fila, column=filas_index)

                self.apply_style(cell, style=style, font_color=font_color)
        else:  # Con reset index
            if num_columns != 0:
                for columna in range(1, num_index):
                    cell = sheet.cell(row=2, column=columna)
                    self.apply_style(cell, style=style, font_color=font_color)

            for columna in range(num_columns + 1, (num_columns + num_index) + 1):
                cell = sheet.cell(row=1, column=columna)
                self.apply_style(cell, style=style, font_color=font_color)

        # Formato por valores
        if pivot_table_config.reset_index:
            start_column = 2
        if filas_columns > 1:
            start_row: int = filas_columns + 2
        else:
            start_row: int = 2

        if pivot_table_config.reset_index:
            start_column = num_columns + num_index + 1

            if num_columns == 0:
                start_column += 1
                columnas_multindex += 1

            for columna in range(2, columnas_multindex + 1):
                for fila in range(2, 2 + len(df_pivot)):
                    cell = sheet.cell(row=fila, column=columna)
                    cell.font = self.bold_font
                    cell.border = self.black_border
        else:
            start_column: int = filas_index + 1

        conteo = self.contar_columnas(tuplas=df_pivot.columns)

        agg_funcs: dict = {}

        if filter_pivot_custom is None:
            agg_funcs = pivot_table_config.aggfunc
        else:
            agg_funcs = filter_pivot_custom.agg_func_params

        for key, func in agg_funcs.items():
            if func == "sum" or func == "mean":
                style = "$#,##0.00"
            else:
                style = "#,##0"

            diff: int = int(conteo[key])

            self._logger.info(f"key: {key}, diff: {diff}")

            if diff == 0:
                diff = 1

            for row in range(start_row, start_row + len(df_pivot)):
                for column in range(start_column, start_column + diff):
                    cell = sheet.cell(row=row, column=column)
                    cell.number_format = style
                    cell.font = self.bold_font
                    cell.border = self.black_border

            start_column += diff

        self._logger.info("Formateo de la tabla dinamica")

    def apply_style(self, cell, style, font_color) -> None:
        """
        Aplica estilo de celda en Excel.

        Args:
            cell (Cell): Celda a la que se aplicará el estilo.
            style (str): Código de color hexadecimal para el fondo de la celda.

        Returns:
            None
        """
        cell.fill = PatternFill(start_color=style, end_color=style, fill_type="solid")

        cell.font = Font(color=font_color, bold=True)

        cell.alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Side(border_style="thin", color="000000")
        cell.border = Border(
            left=thin_border, right=thin_border, top=thin_border, bottom=thin_border
        )

    def contar_columnas(
        self, tuplas: list[tuple]
    ) -> Counter:  # Ensure tuplas is a list of tuples
        """
        Cuenta las repeticiones del último elemento de cada tupla.

        Args:
            tuplas (list[tuple]): Lista de tuplas.

        Returns:
            Counter: Conteo de repeticiones del último elemento.
        """
        primeros = [t[0] for t in tuplas]  # Extraer el último elemento de cada tupla
        return Counter(primeros)

    def _get_aggfunc(self, pivot_table_config: PivotTable) -> PivotOptions | None:
        """Obtiene la lista aggfunc correspondinete de pivot_template"""
        pivot_template: PivotCustom | None = self._pivot_template_dao.get_sync(
            project_id=self.project_id
        )

        if not pivot_template:
            return None

        for pivot_config in pivot_template.pivot_tables:
            if pivot_config.pivot_id == pivot_table_config.id:
                self._logger.info(f"pivot_config: {pivot_template}")
                return pivot_config
        return None
