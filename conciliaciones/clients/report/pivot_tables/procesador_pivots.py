from copy import copy
from io import BytesIO

import pandas as pd
from k_link.db.core import ObjectId
from k_link.db.daos import (
    PivotCustomDAO,
    PivotTableDAO,
)
from k_link.db.models import PivotTable
from k_link.db.models.pivot_customs import PivotCustom
from k_link.extensions.conciliation_type import ConciliationType
from k_link.extensions.pipeline import PivotOptions
from k_link.utils.files.abc_datasource import DataFrame
from k_link.utils.pydantic_types.date import Date
from loggerk import LoggerK
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

from conciliaciones.clients.report.pivot_tables.pivot_handler import PivotHandler
from conciliaciones.clients.report.utils.report_data_handler import ReportDataHandler
from conciliaciones.utils.completion_handler.airflow_contex_exception import (
    AirflowContexException,
)
from conciliaciones.utils.data.normalize import normalize_date
from conciliaciones.utils.redis.redis_keys import RedisKeys
from conciliaciones.utils.redis.redis_storage import RedisStorage

PIVOT_CUSTOM_DAO = PivotCustomDAO()


class ProcesadorPivots:
    def __init__(
        self,
        run_id: str,
        project_id_str: str,
        month: int,
        year: int,
        conciliation_type: ConciliationType,
    ) -> None:
        self._logger = LoggerK(self.__class__.__name__)
        self.run_id = run_id
        self.project_id_str = project_id_str
        self.project_id = ObjectId(project_id_str)
        self.month = month
        self.year = year

        self._pivot_tables = PivotTableDAO()
        self.redis = RedisStorage()
        self._reporte_data = ReportDataHandler(
            run_id=run_id,
            project_id_str=project_id_str,
            month=month,
            year=year,
            conciliation_type=conciliation_type,
        )

        self._redis_key = RedisKeys(
            run_id=run_id,
            project_id_str=project_id_str,
            month=month,
            year=year,
            conciliation_type=conciliation_type,
        )

        self._airflow_fail_exception = AirflowContexException(
            year=year,
            month=month,
            project_id=project_id_str,
            run_id=run_id,
            conciliation_type=conciliation_type,
        )

        self._pivot_table_dao = PivotTableDAO()
        self._pivot_custom: PivotCustom | None = PIVOT_CUSTOM_DAO.get_sync(
            project_id=self.project_id
        )

    async def create_pivot_tables_sheets(self):
        """
        * Orquesta la creación de las hojas de tablas dinámicas y su combinación
        con el archivo Excel principal.
        * Guarda el resultado en Redis.
        """
        if self._pivot_custom is None:
            return

        if not self._pivot_custom.pivot_tables:
            self._logger.info("No hay pivot tables en el cátalogo customs")
            return

        today: Date = normalize_date(self.year, self.month)
        self._logger.info(f"ProjectId: {self.project_id}   Fecha: {today} ")

        result: (
            None | tuple[BytesIO, dict[str, tuple[DataFrame, PivotTable]]]
        ) = await self._generate_pivot_tables_buffer()

        if result is None:
            self._logger.info(
                "No hay configuración de tablas dinámicas, se saltará esta etapa."
            )
            return

        pivot_buffer, names_pivot_table = result

        excel_buffer: BytesIO | None = self._get_excel_buffer()

        if excel_buffer is None:
            self._airflow_fail_exception.handle_and_store_exception(
                f"No se encontró el buffer de Excel para el proyecto: {self.project_id} en la hoja de Tablas Dinámicas."
            )
            return

        combined_buffer: BytesIO = self._combine_workbooks(
            excel_buffer=excel_buffer,
            pivot_buffer=pivot_buffer,
            names_pivot_table=names_pivot_table,
        )

        format_config_key = self._redis_key.get_pivot_table_format_config_key()
        self.redis.set(key=format_config_key, value=names_pivot_table)
        self.redis.set(
            key=self._redis_key.get_excel_buffer_key(), value=combined_buffer
        )
        self._logger.info("Buffer de Excel combinado guardado exitosamente en Redis")

    def _get_excel_buffer(self) -> BytesIO | None:
        """
        * Recupera el buffer del archivo Excel principal desde Redis.
        ! Lanza un error si el buffer no existe.
        ? Retorna: BytesIO con el contenido del Excel.
        """

        buffer_key: str = self._redis_key.get_excel_buffer_key()
        buffer: BytesIO | None = self.redis.get(key=buffer_key, object_type=BytesIO)

        if buffer is None:
            self._airflow_fail_exception.handle_and_store_exception(
                f"No se encontró el buffer de Excel para el proyecto: {self.project_id} en la hoja de Tablas Dinámicas."
            )

        return buffer

    async def _generate_pivot_tables_buffer(
        self,
    ) -> None | tuple[BytesIO, dict[str, tuple[DataFrame, PivotTable]]]:
        """
        * Genera un buffer Excel con las hojas de tablas dinámicas según la configuración.
        * Devuelve el buffer y un diccionario con los nombres y configuraciones de las tablas.
        ? Retorna: Tuple con BytesIO y dict de nombres/configuración.
        """

        pivot_tables: list[PivotOptions] = self._reporte_data.get_pivot_tables_config()

        if not pivot_tables:
            return None

        pivot_tables_config: list[PivotTable] = await self._fetch_pivot_tables_config(
            pivot_tables
        )
        names_pivot_table: dict[str, tuple[DataFrame, PivotTable]] = {}

        try:
            valid_sheets: list[tuple[DataFrame, PivotTable]] = []

            for pivot_table, pivot_option in zip(
                pivot_tables_config, pivot_tables, strict=False
            ):
                df_pivot: DataFrame | None = await self._get_and_transform_data(
                    pivot_table=pivot_table, pivot_options=pivot_option
                )

                if df_pivot is None or df_pivot.empty:
                    self._logger.info(
                        f"Dataframe vacío, pivot table id: {pivot_option.pivot_id}"
                    )
                    continue

                valid_sheets.append((df_pivot, pivot_table))
                names_pivot_table[pivot_table.nombre] = (df_pivot, pivot_table)

            # Si no hay hojas válidas, retornar None
            if not valid_sheets:
                self._logger.info("No hay tablas dinámicas válidas para crear")
                return None

            pivot_buffer = BytesIO()
            with pd.ExcelWriter(pivot_buffer, engine="xlsxwriter") as writer:  # type: ignore
                for df_pivot, pivot_table in valid_sheets:
                    df_pivot.to_excel(writer, sheet_name=pivot_table.nombre)
        except Exception as e:
            self._airflow_fail_exception.handle_and_store_exception(
                f"Error al generar las tablas dinámicas: {e}"
            )
            return None

        pivot_buffer.seek(0)
        return pivot_buffer, names_pivot_table

    async def _get_and_transform_data(
        self, pivot_table: PivotTable, pivot_options: PivotOptions
    ) -> DataFrame | None:
        """
        * Obtiene los datos y aplica transformaciones necesarias (agrupaciones, pivoteo)
        para una tabla dinámica.
        TODO @param pivot_table: Configuración de la tabla dinámica.
        ? Retorna: DataFrame transformado o None si falla.
        """

        pivot_config_catalog: PivotTable | None = self._pivot_table_dao.get_by_id_sync(
            item_id=pivot_options.pivot_id
        )

        if pivot_config_catalog is None:
            return DataFrame()

        df: DataFrame = await self._reporte_data.get_pivot_table_data(
            origen_df=pivot_table.origen,
            pivot_name=pivot_config_catalog.nombre,
            funcion_sheet=pivot_table.funcion_sheet,
        )

        if pivot_table.group_by:
            df = self._reporte_data.group_by_apply(
                df=df,
                group_by=(
                    pivot_options.group_by_params
                    if pivot_options.group_by_params
                    else pivot_table.group_by
                ),
            )

        pivot_handler = PivotHandler(
            df=df,
            pipeline=False,
            pivot_options=pivot_options,
            redis_key=self._redis_key,
        )

        pivot_name, df_pivot = pivot_handler.create_pivot_tables()

        return df_pivot

    def _combine_workbooks(
        self, excel_buffer: BytesIO, pivot_buffer: BytesIO, names_pivot_table: dict
    ) -> BytesIO:
        """
        * Combina el archivo Excel principal con las hojas de tablas dinámicas generadas.
        * Copia hojas, estilos y formatos especiales.
        TODO @param excel_buffer: Buffer del Excel principal.
        TODO @param pivot_buffer: Buffer con hojas de tablas dinámicas.
        TODO @param names_pivot_table: Diccionario de nombres/configuración.
        ? Retorna: Buffer combinado BytesIO.
        """

        existing_workbook: Workbook = load_workbook(excel_buffer)
        pivot_workbook: Workbook = load_workbook(pivot_buffer)

        for sheet_name in pivot_workbook.sheetnames:
            self._copy_sheet(
                source=pivot_workbook[sheet_name],  # type: ignore
                target_wb=existing_workbook,
                sheet_name=sheet_name,
                names_pivot_table=names_pivot_table,
            )

        combined_buffer = BytesIO()
        existing_workbook.save(combined_buffer)
        combined_buffer.seek(0)
        return combined_buffer

    async def _fetch_pivot_tables_config(
        self, pivot_tables: list[PivotOptions]
    ) -> list[PivotTable]:
        """
        * Obtiene la configuración de las tablas dinámicas desde la base de datos.
        * Filtra aquellas que no existen y registra advertencias.
        TODO @param pivot_tables: Lista de opciones de tablas dinámicas.
        ? Retorna: Lista de configuraciones PivotTable.
        """

        config: list[PivotTable] = []

        for pivot in pivot_tables:
            pivot_mongo: PivotTable | None = await self._pivot_tables.get_by_id(
                pivot.pivot_id
            )

            if pivot_mongo is None:
                self._logger.warning(
                    f"Pivot table with ID {pivot.pivot_id} not found in database."
                )
                continue

            self._logger.info(f"Pivot table: {pivot_mongo.nombre}")
            config.append(pivot_mongo)

        return config

    def _copy_sheet(
        self,
        source: Worksheet,
        target_wb: Workbook,
        sheet_name: str,
        names_pivot_table: dict,
    ):
        """
        * Copia una hoja de Excel de un workbook a otro, incluyendo estilos
        y celdas combinadas.
        TODO @param source: Hoja origen.
        TODO @param target_wb: Workbook destino.
        TODO @param sheet_name: Nombre de la hoja.
        TODO @param names_pivot_table: Diccionario de nombres/configuración.
        """

        target = target_wb.create_sheet(title=sheet_name)
        for row in source.iter_rows():
            for cell in row:
                target_cell = target.cell(  # type: ignore
                    row=cell.row, column=cell.column, value=cell.value
                )
                if cell.has_style:
                    target_cell.font = copy(cell.font)
                    target_cell.border = copy(cell.border)
                    target_cell.fill = copy(cell.fill)
                    target_cell.number_format = cell.number_format
                    target_cell.protection = copy(cell.protection)
                    target_cell.alignment = copy(cell.alignment)

        for merged_range in source.merged_cells.ranges:  # type: ignore
            try:
                target.merge_cells(str(merged_range))  # type: ignore
            except Exception as e:
                self._logger.warning(
                    f"No se pudo copiar merged range {merged_range}: {e}"
                )

        self._format_multiindex_if_needed(
            sheet_name=sheet_name,
            target_sheet=target,
            names_pivot_table=names_pivot_table,
        )
        self._copy_dimensions(source_sheet=source, target_sheet=target)

    def _format_multiindex_if_needed(
        self, sheet_name: str, target_sheet, names_pivot_table: dict
    ):
        """
        * Aplica formato especial para MultiIndex en la primera columna si
        la configuración lo requiere.
        TODO @param sheet_name: Nombre de la hoja.
        TODO @param target_sheet: Hoja destino.
        TODO @param names_pivot_table: Diccionario de nombres/configuración.
        """

        pivot_name = sheet_name.replace("Resumen ", "")
        if pivot_name not in names_pivot_table:
            return

        _, config = names_pivot_table[pivot_name]
        if config.reset_index:
            return

        self._logger.info(f"Aplicando formato MultiIndex para {sheet_name}")
        if target_sheet.max_row <= 1:
            return

        current_value, start_row = None, 2
        for row_idx in range(2, target_sheet.max_row + 2):
            cell_value = (
                target_sheet.cell(row=row_idx, column=1).value
                if row_idx <= target_sheet.max_row
                else None
            )
            if cell_value != current_value:
                if current_value is not None and start_row < row_idx - 1:
                    try:
                        target_sheet.merge_cells(
                            start_row=start_row,
                            start_column=1,
                            end_row=row_idx - 1,
                            end_column=1,
                        )
                        cell = target_sheet.cell(row=start_row, column=1)
                        cell.alignment = Alignment(vertical="center", horizontal="left")
                        self._logger.info(
                            f"Merged filas {start_row} a {row_idx - 1} para valor: {current_value}"
                        )
                    except Exception as e:
                        self._logger.warning(
                            f"No se pudo hacer merge filas {start_row}-{row_idx - 1}: {e}"
                        )
                current_value, start_row = cell_value, row_idx

    def _copy_dimensions(self, source_sheet: Worksheet, target_sheet):
        """
        * Copia las dimensiones (ancho de columnas y alto de filas) de la hoja
        origen a la destino.
        TODO @param source_sheet: Hoja origen.
        TODO @param target_sheet: Hoja destino.
        """

        for col in source_sheet.column_dimensions:
            if source_sheet.column_dimensions[col].width:
                target_sheet.column_dimensions[
                    col
                ].width = source_sheet.column_dimensions[col].width
        for row in source_sheet.row_dimensions:
            if source_sheet.row_dimensions[row].height:
                target_sheet.row_dimensions[row].height = source_sheet.row_dimensions[
                    row
                ].height
