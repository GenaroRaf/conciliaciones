from io import BytesIO

from k_link.db.core import ObjectId
from k_link.db.daos import ERPFilesDAO, LinkServicesDAO, ProjectDAO, ReportCatalogDAO
from k_link.extensions.conciliation_type import ConciliationType
from k_link.extensions.datasources.data_sources import DataSources
from loggerk import LoggerK
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from conciliaciones.clients.report.utils.report_data_handler import ReportDataHandler
from conciliaciones.utils.completion_handler.airflow_contex_exception import (
    AirflowContexException,
)
from conciliaciones.utils.data.normalize import normalize_date
from conciliaciones.utils.redis.redis_keys import RedisKeys
from conciliaciones.utils.redis.redis_storage import RedisStorage


class ResumenSheet:
    """
    * Clase para crear la hoja de resumen del reporte.
    * Incluye información general del proyecto, archivos conciliados, tipo de reporte y periodo.
    """

    def __init__(
        self,
        run_id: str,
        project_id_str: str,
        month: int,
        year: int,
        conciliation_type: ConciliationType,
    ) -> None:
        """
        Inicializa la clase con los parámetros necesarios para generar el resumen.

        @param run_id: Identificador de ejecución del reporte.
        @param project_id_str: ID del proyecto como string.
        @param month: Mes del periodo conciliado.
        @param year: Año del periodo conciliado.
        @param conciliation_type: Tipo de conciliación (ERP, SAT, etc).
        @param tipo_reporte: Tipo de reporte generado (ej. global, detallado).
        """
        self._logger = LoggerK(self.__class__.__name__)

        self._project_dao = ProjectDAO()
        self._erp_files_dao = ERPFilesDAO()
        self._link_services_dao = LinkServicesDAO()
        self._report_catalog_dao = ReportCatalogDAO()

        self._project_id = ObjectId(project_id_str)
        self.redis = RedisStorage()
        self.month: int = month
        self.year: int = year
        self._conciliation_type: ConciliationType = conciliation_type

        self._reporte_data = ReportDataHandler(
            run_id=run_id,
            project_id_str=project_id_str,
            month=month,
            year=year,
            conciliation_type=conciliation_type,
        )

        self._redis_key = RedisKeys(
            run_id=run_id,
            project_id_str=project_id_str,
            month=month,
            year=year,
            conciliation_type=conciliation_type,
        )
        self._airflow_fail_exception = AirflowContexException(
            year=year,
            month=month,
            project_id=project_id_str,
            run_id=run_id,
            conciliation_type=conciliation_type,
        )

    async def create_sheet_resumen(self) -> None:
        """
        Crea la hoja de resumen y la guarda en Redis como un archivo Excel en memoria.
        """

        today = normalize_date(self.year, self.month)
        self._logger.info(f"ProjectId: {self._project_id}   Fecha: {today}")

        workbook: Workbook = Workbook()
        worksheet: Worksheet = workbook.active  # type: ignore
        worksheet.title = "Resumen"

        await self._crear_info_resumen(worksheet)

        new_excel_buffer: BytesIO = BytesIO()
        workbook.save(new_excel_buffer)
        new_excel_buffer.seek(0)

        excel_buffer_key: str = self._redis_key.get_excel_buffer_key()
        self.redis.set(key=excel_buffer_key, value=new_excel_buffer)

    async def _crear_info_resumen(self, worksheet: Worksheet) -> None:
        """
        Escribe toda la información del resumen del proyecto sobre la hoja proporcionada.

        @param worksheet: Objeto Worksheet de openpyxl donde se escribirá la información.
        """

        proyecto = await self._project_dao.get_by_id(item_id=self._project_id)
        if not proyecto:
            self._airflow_fail_exception.handle_and_store_exception(
                f"Proyecto con ID: {self._project_id} no encontrado."
            )

        # Definición de estilos
        header_font: Font = Font(bold=True, size=14, color="FFFFFF")
        subheader_font: Font = Font(bold=True, size=12, color="000000")
        data_font: Font = Font(size=12, color="000000")

        header_fill: PatternFill = PatternFill("solid", fgColor="065E8E")  # Azul
        subheader_fill: PatternFill = PatternFill(
            "solid", fgColor="36C18C"
        )  # Verde menta
        data_fill: PatternFill = PatternFill("solid", fgColor="FFFFFF")  # Blanco

        border: Border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        align_center: Alignment = Alignment(horizontal="center")

        def style_cell(
            cell,
            font: Font,
            fill: PatternFill,
            border: Border = border,
            align: Alignment = align_center,
        ) -> None:
            """
            Aplica estilos de fuente, relleno, borde y alineación a una celda.

            @param cell: Celda de openpyxl.
            @param font: Estilo de fuente.
            @param fill: Relleno de fondo.
            @param border: Borde aplicado.
            @param align: Alineación del texto.
            """

            cell.font = font
            cell.fill = fill
            cell.border = border
            cell.alignment = align

        # Encabezado del proyecto
        worksheet.merge_cells("B2:G2")
        style_cell(worksheet["B2"], header_font, header_fill)
        worksheet["B2"].value = "Proyecto"

        worksheet.merge_cells("B3:G3")
        style_cell(worksheet["B3"], data_font, data_fill)
        worksheet["B3"].value = proyecto.name

        # Archivos conciliados
        worksheet["B5"].value = "Archivos conciliados"
        style_cell(
            worksheet["B5"],
            subheader_font,
            subheader_fill,
            align=Alignment(horizontal="left"),
        )

        row: int = 6

        if proyecto.project_type == "Fiscal":
            link_services: list[dict] = await self._link_services_dao.agregate(
                [
                    {"$match": {"project_id": self._project_id}},
                    {
                        "$project": {
                            "_id": 0,
                            "report_id": "$report_config.report_type.report_id",
                        }
                    },
                ]
            )

            report_id: ObjectId = ObjectId(link_services[0]["report_id"])

            report_type: dict | None = await self._report_catalog_dao.get_no_pydantic(
                filters={"_id": report_id}, projection=["name"]
            )

            if report_type is None:
                self._airflow_fail_exception.handle_and_store_exception(
                    f"No se encontró el tipo de reporte con ID: {report_id}"
                )

            report_name: str = report_type["name"]

            cell = worksheet["B6"]
            cell.value = report_name
            style_cell(cell, data_font, data_fill, align=Alignment(horizontal="left"))
            row = 7
        else:
            erp_file = await self._erp_files_dao.get(project_id=self._project_id)
            if not erp_file:
                self._airflow_fail_exception.handle_and_store_exception(
                    f"No hay configuración de ERPFiles para el proyecto: {self._project_id}"
                )

            data_sources: list[DataSources] = erp_file.data_sources
            for ds in data_sources:
                cell = worksheet[f"B{row}"]
                cell.value = ds.config.datasource_name
                style_cell(
                    cell, data_font, data_fill, align=Alignment(horizontal="left")
                )
                row += 1

        exiting_report = await self._reporte_data.get_info_metadata_report()
        execution_date = exiting_report.creation_date

        if execution_date is None:
            execution_date_str = "N/A"

        # Tipo de reporte
        worksheet[f"B{row}"].value = "Reporte Kuantik"
        style_cell(
            worksheet[f"B{row}"],
            subheader_font,
            subheader_fill,
            align=Alignment(horizontal="left"),
        )
        row += 1

        worksheet[f"B{row}"].value = proyecto.project_type
        style_cell(
            worksheet[f"B{row}"],
            data_font,
            data_fill,
            align=Alignment(horizontal="left"),
        )
        row += 1

        # Periodo conciliado
        worksheet[f"B{row}"].value = "Periodo conciliado"
        style_cell(
            worksheet[f"B{row}"],
            subheader_font,
            subheader_fill,
            align=Alignment(horizontal="left"),
        )
        row += 1

        worksheet[f"B{row}"].value = f"{self.month}/{self.year}"
        style_cell(
            worksheet[f"B{row}"],
            data_font,
            data_fill,
            align=Alignment(horizontal="left"),
        )
        row += 1

        # Fecha de ejecución
        worksheet[f"B{row}"].value = "Fecha de ejecución"
        style_cell(
            worksheet[f"B{row}"],
            subheader_font,
            subheader_fill,
            align=Alignment(horizontal="left"),
        )
        row += 1

        if execution_date is not None:
            worksheet[f"B{row}"].value = execution_date.strftime("%d/%m/%Y %H:%M:%S")

        style_cell(
            worksheet[f"B{row}"],
            data_font,
            data_fill,
            align=Alignment(horizontal="left"),
        )
        row += 3

        # Ajuste de ancho de columnas
        for col in range(1, 27):  # Columnas A-Z
            col_letter = get_column_letter(col)
            worksheet.column_dimensions[col_letter].width = 20
