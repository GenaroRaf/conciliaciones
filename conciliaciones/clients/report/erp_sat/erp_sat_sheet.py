import datetime
from io import BytesIO

import numpy as np
import pandas as pd
from k_link.db.core import ObjectId
from k_link.db.daos import ERPFilesDAO, ProjectDAO
from k_link.db.models.report_catalog import ReportCatalog
from k_link.extensions.conciliation_type import ConciliationType
from k_link.extensions.report_config import OrigenColumna, TipoDato
from k_link.extensions.report_config.header_config import HeaderConfig
from k_link.utils.files.abc_datasource import DataFrame
from loggerk import LoggerK
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from conciliaciones.clients.report.styles.excel_styles import OpenpyxlStylesExcel
from conciliaciones.clients.report.styles.report_styles import ReportStyles
from conciliaciones.clients.report.styles.xlsxwritter_excel_styles import HeaderStyles
from conciliaciones.clients.report.utils.report_data_handler import ReportDataHandler
from conciliaciones.utils.completion_handler.airflow_contex_exception import (
    AirflowContexException,
)
from conciliaciones.utils.data.normalize import normalize_date
from conciliaciones.utils.redis.redis_keys import RedisKeys
from conciliaciones.utils.redis.redis_storage import RedisStorage


class ERPSATSheet:
    """
    * Clase encargada de crear la hoja combinada ERP + SAT en el archivo de Excel.
    * Integra los datos conciliados y les aplica formato visual.
    """

    _sheet_name: str = "ERP + SAT"
    _styles: OpenpyxlStylesExcel

    def __init__(
        self,
        run_id: str,
        project_id_str: str,
        month: int,
        year: int,
        conciliation_type: ConciliationType,
    ) -> None:
        """
        * Inicializa la clase con los datos necesarios para generar la hoja ERP + SAT.

        TODO @param run_id: ID de ejecución del reporte.
        TODO @param project_id_str: ID del proyecto en formato string.
        TODO @param month: Mes del periodo conciliado.
        TODO @param year: Año del periodo conciliado.
        TODO @param conciliation_type: Tipo de conciliación.
        TODO @param tipo_reporte: Tipo de reporte a generar.
        """

        self._logger = LoggerK(self.__class__.__name__)

        self._project_dao = ProjectDAO()
        self._erp_files_dao = ERPFilesDAO()

        self.project_id = ObjectId(project_id_str)
        self.redis = RedisStorage()
        self.month = month
        self.year = year
        self._conciliation_type = conciliation_type

        self._reporte_data = ReportDataHandler(
            run_id=run_id,
            project_id_str=project_id_str,
            month=month,
            year=year,
            conciliation_type=conciliation_type,
        )
        self._redis_key = RedisKeys(
            run_id=run_id,
            project_id_str=project_id_str,
            month=month,
            year=year,
            conciliation_type=conciliation_type,
        )
        self._airflow_fail_exception = AirflowContexException(
            year=year,
            month=month,
            project_id=project_id_str,
            run_id=run_id,
            conciliation_type=conciliation_type,
        )
        self._report_styles = ReportStyles(
            run_id=run_id,
            project_id_str=project_id_str,
            month=month,
            year=year,
            conciliation_type=conciliation_type,
        )

    async def create_erp_sat_sheet(self) -> None:
        """
        * Crea la hoja ERP + SAT en el archivo de Excel almacenado en Redis.
        * Inserta los datos y aplica estilos personalizados.
        """

        today = normalize_date(self.year, self.month)
        self._logger.info(f"ProjectId: {self.project_id}   Fecha: {today}")

        excel_buffer_key: str = self._redis_key.get_excel_buffer_key()
        excel_buffer: BytesIO | None = self.redis.get(
            key=excel_buffer_key, object_type=BytesIO
        )

        if excel_buffer is None:
            self._airflow_fail_exception.handle_and_store_exception(
                f"No se encontró el buffer de Excel para el proyecto: {self.project_id} en la hoja ERP + SAT."
            )

        excel_buffer.seek(0)
        workbook: Workbook = load_workbook(excel_buffer)

        # Verifica si la hoja ya existe y la elimina si es necesario
        if self._sheet_name in workbook.sheetnames:
            del workbook[self._sheet_name]

        worksheet: Worksheet = workbook.create_sheet(title=self._sheet_name)

        self._styles = OpenpyxlStylesExcel()

        df_erp_sat: DataFrame = await self._reporte_data.get_erp_sat_data()
        df_erp_sat = await self._clean_dataframe(df_erp_sat)

        self._logger.info(f"DataFrame ERP + SAT: {df_erp_sat.columns}")

        duplicate_headers = df_erp_sat.columns[df_erp_sat.columns.duplicated()].tolist()

        self._logger.error(f"Headers duplicados en ERP + SAT: {duplicate_headers}")

        df_erp_sat = df_erp_sat.loc[:, ~df_erp_sat.columns.duplicated()]

        # Insertar datos en la hoja, validando si el header ya existe
        for r_idx, row in enumerate(df_erp_sat.to_records(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                worksheet.cell(
                    row=r_idx, column=c_idx, value=self.sanitize_excel_value(value)
                )

        await self._erp_sat_format(worksheet, df_erp_sat)

        # Guardar nuevamente el archivo en Redis
        new_excel_buffer: BytesIO = BytesIO()
        workbook.save(new_excel_buffer)
        new_excel_buffer.seek(0)

        self._logger.info(f"ERP + SAT headers: {df_erp_sat.columns.tolist()}")
        self._logger.info(f"Saving ERP + SAT sheet: {df_erp_sat}")

        self.redis.set(key=excel_buffer_key, value=new_excel_buffer)

    async def _clean_dataframe(self, df: DataFrame) -> DataFrame:
        """
        * Limpia y normaliza columnas monetarias del DataFrame según configuración.

        TODO @param df: DataFrame con datos ERP + SAT.
        TODO @return: DataFrame limpio y listo para exportar.
        """

        headers_validos = self._reporte_data.get_tabla_config()

        for column_config in headers_validos:
            tipo_dato: TipoDato = column_config.configuracion_tipo_dato
            origen_colum: OrigenColumna = column_config.origen

            if self.validate_data_and_clean_data(tipo_dato, origen_colum):
                column_name: str = column_config.nombre
                if column_name in df.columns:
                    df[column_name] = df[column_name].fillna("0.00")

        return df

    async def _erp_sat_format(
        self, worksheet: Worksheet, df_erp_sat: pd.DataFrame
    ) -> None:
        """
        * Aplica formato a la hoja ERP + SAT, incluyendo encabezados estilizados.

        TODO @param worksheet: Hoja de Excel donde se aplicará el formato.
        TODO @param df_erp_sat: DataFrame con los datos a representar.
        """

        report_type: ReportCatalog | None = self._reporte_data.get_report_type()

        if report_type is None:
            self._logger.error("No se encontró el tipo de reporte asociado.")
            return

        headers_erp_sat: list[
            HeaderConfig
        ] = await self._reporte_data.get_headers_erp_sat(df=df_erp_sat)

        # Configuración de encabezados
        for col_num, header_erp_sat in enumerate(headers_erp_sat, start=1):
            # Se recupera diccionario de headers del reporte correspondiente
            headers_report_kreports: dict[
                str, str
            ] = await self._reporte_data.get_headers_report_kreports(
                report_type=report_type
            )

            # Se asigna nombre del diccionario de headers kreports
            if header_erp_sat.nombre in headers_report_kreports:
                header_cell = worksheet.cell(
                    row=1,
                    column=col_num,
                    value=headers_report_kreports[header_erp_sat.nombre],
                )
            else:
                header_cell = worksheet.cell(
                    row=1, column=col_num, value=header_erp_sat.nombre
                )

            # Ajustar ancho de columna a 20
            col_letter = get_column_letter(col_num)
            worksheet.column_dimensions[col_letter].width = 20

            origen_header: str = header_erp_sat.origen.value
            header_style_enum: HeaderStyles = HeaderStyles[origen_header]
            header_cell.style = self._styles.header_styles[header_style_enum]

        for column_num, column_config in enumerate(headers_erp_sat, start=1):
            column_name: str = column_config.nombre
            tipo_dato: TipoDato = column_config.configuracion_tipo_dato
            origen_filas: OrigenColumna = column_config.origen

            self._report_styles.styles_formart(
                worksheet=worksheet,
                df=df_erp_sat,
                column_name=column_name,
                column_num=column_num,
                tipo_dato=tipo_dato,
                origen=origen_filas,
            )

    def sanitize_excel_value(self, val):
        """
        * Convierte valores incompatibles con openpyxl a tipos válidos.
        * Soporta numpy.datetime64, numpy.generic (int, float, bool), etc.
        """
        if isinstance(val, np.datetime64):
            return val.astype("M8[ms]").astype(datetime.datetime)
        if isinstance(val, np.generic):
            return val.item()
        return val

    def validate_data_and_clean_data(
        self, tipo_dato: TipoDato, origin_column: OrigenColumna
    ) -> bool:
        """
        * Valida si una columna debe ser limpiada con valor por defecto.

        TODO @param tipo_dato: Tipo de dato de la columna (ej. MONEDA).
        TODO @param origin_column: Origen de la columna (ej. CLEAN_DATA).
        TODO @return: True si se debe limpiar, False si no.
        """

        return (
            tipo_dato == TipoDato.MONEDA and origin_column == OrigenColumna.CLEAN_DATA
        )
