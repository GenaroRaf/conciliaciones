from io import BytesIO

from k_link.db.core import ObjectId
from k_link.db.daos import ERPFilesDAO, ProjectDAO
from k_link.extensions.conciliation_type import ConciliationType
from k_link.utils.files.abc_datasource import DataFrame
from loggerk import LoggerK
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from conciliaciones.clients.report.styles.xlsxwritter_excel_styles import (
    DataStyles,
    HeaderStyles,
)
from conciliaciones.clients.report.utils.report_data_handler import ReportDataHandler
from conciliaciones.utils.completion_handler.airflow_contex_exception import (
    AirflowContexException,
)
from conciliaciones.utils.data.normalize import normalize_date
from conciliaciones.utils.redis.redis_keys import RedisKeys
from conciliaciones.utils.redis.redis_storage import RedisStorage


class MetadataSheets:
    """
    * Clase encargada de crear las hojas de metadatos en el archivo de Excel.
    * Incluye hojas para datos pendientes y cancelados.
    """

    _sheet_pending = "Descarga Pendiente"
    _sheet_cancelled = "Metadata Cancelada"

    def __init__(
        self,
        run_id: str,
        project_id_str: str,
        month: int,
        year: int,
        conciliation_type: ConciliationType,
    ) -> None:
        """
        * Inicializa la clase con los datos necesarios para generar las hojas extras.
        TODO @param run_id: ID de ejecución del reporte.
        TODO @param project_id_str: ID del proyecto en formato string.
        TODO @param month: Mes del periodo conciliado.
        TODO @param year: Año del periodo conciliado.
        TODO @param conciliation_type: Tipo de conciliación.
        TODO @param tipo_reporte: Tipo de reporte a generar.
        """

        self._logger = LoggerK(self.__class__.__name__)

        self._project_dao = ProjectDAO()
        self._erp_files_dao = ERPFilesDAO()

        self.project_id = ObjectId(project_id_str)
        self.redis = RedisStorage()
        self.month = month
        self.year = year
        self._conciliation_type: ConciliationType = conciliation_type

        self._reporte_data = ReportDataHandler(
            run_id=run_id,
            project_id_str=project_id_str,
            month=month,
            year=year,
            conciliation_type=conciliation_type,
        )
        self._redis_key = RedisKeys(
            run_id=run_id,
            project_id_str=project_id_str,
            month=month,
            year=year,
            conciliation_type=conciliation_type,
        )
        self._airflow_fail_exception = AirflowContexException(
            year=year,
            month=month,
            project_id=project_id_str,
            run_id=run_id,
            conciliation_type=conciliation_type,
        )

    async def create_metadata_sheets(self):
        """
        * Crea las hojas de metadatos en el archivo de Excel almacenado en Redis.
        * Integra los datos de reportes extras configurados y les aplica formato de SAT.
        ! Si el Excel buffer no existe en Redis, se lanza una excepción.
        """

        today = normalize_date(self.year, self.month)
        self._logger.info(f"ProjectId: {self.project_id}   Fecha: {today} ")

        excel_buffer_key: str = self._redis_key.get_excel_buffer_key()
        excel_buffer: BytesIO | None = self.redis.get(
            key=excel_buffer_key, object_type=BytesIO
        )

        if excel_buffer is None:
            self._airflow_fail_exception.handle_and_store_exception(
                f"No se encontró el buffer de Excel para el proyecto: {self.project_id} en la hoja Metadata."
            )

        excel_buffer.seek(0)
        workbook: Workbook = load_workbook(excel_buffer)

        df_pending: DataFrame = await self._reporte_data.get_pending_data()
        df_cancelled: DataFrame = await self._reporte_data.get_cancelled_data()

        df_metadata: list[tuple[DataFrame, str]] = [
            (df_pending, self._sheet_pending),
            (df_cancelled, self._sheet_cancelled),
        ]

        for df, sheet_name in df_metadata:
            if df.empty:
                self._logger.info(f"No hay datos para la hoja: {sheet_name}")
                continue

            # Verifica si la hoja ya existe y la elimina si es necesario
            if sheet_name in workbook.sheetnames:
                del workbook[sheet_name]

            worksheet = workbook.create_sheet(title=sheet_name)

            self._logger.info(f"Procesando hoja: {sheet_name}")

            self._logger.info(f"DataFrame {sheet_name}: {df.columns}")

            self._logger.info(f"Formato de hoja: {sheet_name}")

            # Escribir los encabezados y datos manualmente
            for r_idx, row in enumerate(df.to_records(index=False), start=2):
                for c_idx, value in enumerate(row, start=1):
                    # Convertir listas, dicts o sets a string legible
                    if isinstance(value, list):
                        value = ", ".join(
                            (
                                "; ".join(f"{k}: {v}" for k, v in item.items())
                                if isinstance(item, dict)
                                else str(item)
                            )
                            for item in value
                        )
                    elif isinstance(value, dict):
                        value = "; ".join(f"{k}: {v}" for k, v in value.items())
                    elif isinstance(value, set):
                        value = ", ".join(str(v) for v in value)

                    worksheet.cell(row=r_idx, column=c_idx, value=value)

            await self.format_metadata(worksheet, df)

            self._logger.info(f"Hoja: {sheet_name} {df}")

        # Se guarda el excel buffer
        self._logger.info(
            f"Guardando buffer de Excel del {self._sheet_pending} y {self._sheet_cancelled} en Redis: {excel_buffer}"
        )

        new_excel_buffer = BytesIO()
        workbook.save(new_excel_buffer)
        new_excel_buffer.seek(0)

        # Guarda el nuevo buffer en Redis
        self.redis.set(key=excel_buffer_key, value=new_excel_buffer)

    async def format_metadata(self, worksheet: Worksheet, df: DataFrame) -> None:
        """
        * Aplica formato a la hoja de metadatos, incluyendo encabezados estilizados.
        TODO @param worksheet: Hoja de Excel donde se aplicará el formato.
        TODO @param df: DataFrame con los datos a representar.
        """

        # Configuración de los encabezados
        for col_idx, col_name in enumerate(df.columns, start=1):
            header_style = HeaderStyles.SAT.value
            font_color = header_style["font_color"]
            bg_color = header_style["bg_color"]

            cell = worksheet.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=False, color=font_color)
            cell.fill = PatternFill(
                start_color=bg_color, end_color=bg_color, fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="left")
            worksheet.column_dimensions[cell.column_letter].width = 20

        # Opcional: aplicar formato a columnas específicas (ejemplo: "Total")
        if "Total" in df.columns:
            total_col_idx = list(df.columns).index("Total") + 1
            for row in worksheet.iter_rows(
                min_row=2,
                min_col=total_col_idx,
                max_col=total_col_idx,
                max_row=worksheet.max_row,
            ):
                for cell in row:
                    data_style = DataStyles["MONEDA"].value
                    cell.number_format = data_style.get("number_format", "#,##0.00")
                    cell.alignment = Alignment(horizontal="right")
