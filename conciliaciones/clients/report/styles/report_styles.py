from datetime import datetime

import pandas as pd
from k_link.db.core import ObjectId
from k_link.db.daos import ERPFilesDAO, ProjectDAO, ReportCatalogDAO
from k_link.db.models import ReportCatalog
from k_link.extensions.conciliation_type import ConciliationType
from k_link.extensions.report_config import (
    DataTypeReport,
    HeaderConfig,
    OrigenColumna,
    TipoDato,
)
from k_link.utils.files.abc_datasource import DataFrame
from loggerk import LoggerK
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from conciliaciones.clients.report.styles.excel_styles import OpenpyxlStylesExcel
from conciliaciones.clients.report.styles.xlsxwritter_excel_styles import (
    DataStyles,
    HeaderStyles,
)
from conciliaciones.clients.report.utils.report_data_handler import ReportDataHandler
from conciliaciones.utils.completion_handler.airflow_contex_exception import (
    AirflowContexException,
)
from conciliaciones.utils.redis.redis_keys import RedisKeys
from conciliaciones.utils.redis.redis_storage import RedisStorage


class ReportStyles:
    def __init__(
        self,
        run_id: str,
        project_id_str: str,
        month: int,
        year: int,
        conciliation_type: ConciliationType,
    ):
        self._logger = LoggerK(self.__class__.__name__)
        self._project_dao = ProjectDAO()
        self._erp_files_dao = ERPFilesDAO()
        self._report_catalog_dao = ReportCatalogDAO()

        self.project_id = ObjectId(project_id_str)
        self.redis = RedisStorage()
        self.month = month
        self.year = year
        self._conciliation_type: ConciliationType = conciliation_type

        self._reporte_data = ReportDataHandler(
            run_id=run_id,
            project_id_str=project_id_str,
            month=month,
            year=year,
            conciliation_type=conciliation_type,
        )
        self._redis_key = RedisKeys(
            run_id=run_id,
            project_id_str=project_id_str,
            month=month,
            year=year,
            conciliation_type=conciliation_type,
        )
        self._airflow_fail_exception = AirflowContexException(
            year=year,
            month=month,
            project_id=project_id_str,
            run_id=run_id,
            conciliation_type=conciliation_type,
        )
        self._styles = OpenpyxlStylesExcel()

    def resolve_column_name(self, original_column_name: str, df_columns: list[str]):
        kuantik_column_name = original_column_name + " (K)"

        for column in [original_column_name, kuantik_column_name]:
            if column in df_columns:
                return column
        return None

    def styles_formart(
        self,
        worksheet: Worksheet,
        df: DataFrame,
        column_name,
        origen: OrigenColumna,
        tipo_dato: TipoDato,
        column_num: int,
    ) -> None:
        if self.validate_moneda_and_sat(tipo_dato=tipo_dato, origin_column=origen):
            column_name = self.resolve_column_name(column_name, df.columns.tolist())
            if column_name is None:
                return

            data_style_enum = DataStyles[tipo_dato.value]

            for i, dato in enumerate(df[column_name], start=2):
                if pd.notna(dato) and str(dato).strip() != "":
                    cell = worksheet.cell(row=i, column=column_num, value=dato)
                    cell.style = self._styles.data_styles[data_style_enum]
                    cell.number_format = '"$"#,##0.00'
        elif self.validate_data_and_erp(tipo_dato=tipo_dato, origin_column=origen):
            column_name = self.resolve_column_name(column_name, df.columns.tolist())
            if column_name is None:
                return

            data_style_enum = DataStyles[tipo_dato.value]

            for i, fecha in enumerate(df[column_name], start=2):
                if pd.notna(fecha) and str(fecha).strip() != "":
                    try:
                        fecha = self._parse_fecha(str(fecha))
                    except ValueError as _:
                        continue
                    cell = worksheet.cell(row=i, column=column_num, value=fecha)
                    cell.style = self._styles.data_styles[data_style_enum]
                    cell.number_format = "dd/mm/yyyy"

        elif self.validate_data_and_clean_data(
            tipo_dato=tipo_dato, origin_column=origen
        ):
            column_name = self.resolve_column_name(column_name, df.columns.tolist())
            if column_name is None:
                return

            data_style_enum = DataStyles[tipo_dato.value]

            for i, dato in enumerate(df[column_name], start=2):
                if isinstance(dato, (list, tuple, set)):
                    dato = ", ".join(map(str, dato))
                elif isinstance(dato, dict):
                    dato = str(dato)

                if pd.notna(dato) and str(dato).strip() != "":
                    cell = worksheet.cell(row=i, column=column_num, value=dato)
                    cell.style = self._styles.data_styles[data_style_enum]

    async def headers_sat(
        self,
        worksheet: Worksheet,
        df: DataFrame,
        erp_sat: bool,
        headers_sat_config: list[HeaderConfig],
    ) -> None:
        report_type: ReportCatalog | None = self._reporte_data.get_report_type()

        if not report_type:
            self._airflow_fail_exception.handle_and_store_exception(
                f"No se encontró el report type para el proyecto: {self.project_id}."
            )

        self._logger.error(f"Headers DF: {df.columns.tolist()}")

        n_column = len(self._reporte_data.filter_columns_erp(df)) if erp_sat else 0

        self._logger.info(f"lista de headers config: {headers_sat_config}")

        # Se recupera diccionario de headers del reporte correspondiente
        headers_report_kreports: dict[
            str, str
        ] = await self._reporte_data.get_headers_report_kreports(
            report_type=report_type
        )

        # Configuracion de los headers
        for column_num, header_sat in enumerate(headers_sat_config, start=n_column + 1):
            sat_style = HeaderStyles.SAT.value
            font_color: str = sat_style["font_color"]
            bg_color: str = sat_style["bg_color"]

            # Se asigna nombre del diccionario de headers kreports
            if header_sat.nombre in headers_report_kreports:
                cell = worksheet.cell(
                    row=1,
                    column=column_num,
                    value=headers_report_kreports[header_sat.nombre],
                )
            else:
                cell = worksheet.cell(row=1, column=column_num, value=header_sat.nombre)

            cell.font = Font(bold=False, color=font_color)
            cell.fill = PatternFill(
                start_color=bg_color, end_color=bg_color, fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="left")
            worksheet.column_dimensions[cell.column_letter].width = 20  # type: ignore

        # Extracion del tipo de dato de ReportCatalog
        for column_num, header_sat in enumerate(headers_sat_config, start=1):
            tipo_dato: TipoDato = header_sat.configuracion_tipo_dato

            # Aplicar formatos a las filas
            self.styles_formart(
                worksheet=worksheet,
                df=df,
                column_name=header_sat.nombre,
                origen=OrigenColumna.SAT,
                tipo_dato=tipo_dato,
                column_num=column_num,
            )

    async def headers_extra_sheets(
        self,
        worksheet: Worksheet,
        df: DataFrame,
        report_id: ObjectId,
    ) -> None:
        report_type: ReportCatalog | None = await self._report_catalog_dao.get_by_id(
            item_id=report_id
        )

        if not report_type:
            self._airflow_fail_exception.handle_and_store_exception(
                f"Extra Sheet con el ID {report_id} no encontrado."
            )

        self._logger.error(f"Headers DF: {df.columns.tolist()}")

        headers_report_type = report_type.headers
        self._logger.info(f"Headers extra sheet: {headers_report_type}")
        headers_sat = {}

        for header in df.columns:
            if header in headers_report_type:
                headers_sat[header] = headers_report_type[header]
            else:
                headers_sat[header] = {"type": "TEXTO", "format": None}
                self._logger.error(
                    f"Header '{header}' no encontrado en ReportCatalog, asignando tipo 'TEXTO' por defecto."
                )

        # Se recupera diccionario de headers del reporte correspondiente
        headers_report_kreports: dict[
            str, str
        ] = await self._reporte_data.get_headers_report_kreports(
            report_type=report_type
        )

        # Configuracion de los headers
        for column_num, (key, value) in enumerate(headers_sat.items(), start=1):
            sat_style = HeaderStyles.SAT.value
            font_color = sat_style["font_color"]
            bg_color = sat_style["bg_color"]

            # Se asigna nombre del diccionario de headers kreports
            if key in headers_report_kreports:
                cell = worksheet.cell(
                    row=1, column=column_num, value=headers_report_kreports[key]
                )
            else:
                cell = worksheet.cell(row=1, column=column_num, value=key)  # type: ignore

            cell.font = Font(bold=False, color=font_color)
            cell.fill = PatternFill(
                start_color=bg_color, end_color=bg_color, fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="left")
            worksheet.column_dimensions[cell.column_letter].width = 20  # type: ignore

        # Extracion del tipo de dato de ReportCatalog
        for column_num, (key, value) in enumerate(headers_sat.items(), start=1):
            type_value = value.get("type")
            self._logger.info(f"Column: {key} - Value: {type_value}")
            format_type = value.get("format", None)
            tipo_dato_report: DataTypeReport = self.parse_tipo_dato_report(
                type_value=type_value, format_type=format_type
            )
            tipo_dato: TipoDato = tipo_dato_report.tipo_dato

            # Aplicar formatos a las filas
            self.styles_formart(
                worksheet=worksheet,
                df=df,
                column_name=key,
                origen=OrigenColumna.SAT,
                tipo_dato=tipo_dato,
                column_num=column_num,
            )

    def _parse_fecha(self, fecha_str: str) -> datetime:
        fecha_str = fecha_str.strip()  # Limpieza clave
        formatos = [
            "%Y-%m-%dT%H:%M:%SZ",  # ISO con Z
            "%Y-%m-%dT%H:%M:%S",  # ISO sin Z
            "%Y-%m-%dT%H:%M",  # ISO sin segundos
            "%Y-%m-%d %H:%M:%S",  # ISO con espacio
            "%Y-%m-%d %H:%M",  # ISO con espacio sin segundos
            "%d/%m/%Y %H:%M:%S",  # DD/MM/YYYY HH:MM:SS
            "%d/%m/%Y %H:%M",  # DD/MM/YYYY HH:MM
            "%d/%m/%Y",  # DD/MM/YYYY
            "%d-%m-%Y %H:%M:%S",  # DD-MM-YYYY HH:MM:SS
            "%d-%m-%Y %H:%M",  # DD-MM-YYYY HH:MM
            "%d-%m-%Y",  # DD-MM-YYYY
            "%m/%d/%Y %H:%M:%S",  # MM/DD/YYYY HH:MM:SS
            "%m/%d/%Y %H:%M",  # MM/DD/YYYY HH:MM
            "%m/%d/%Y",  # MM/DD/YYYY
            "%m-%d-%Y %H:%M:%S",  # MM-DD-YYYY HH:MM:SS
            "%m-%d-%Y %H:%M",  # MM-DD-YYYY HH:MM
            "%m-%d-%Y",  # MM-DD-YYYY
            "%Y/%m/%d %H:%M:%S",  # YYYY/MM/DD HH:MM:SS
            "%Y/%m/%d %H:%M",  # YYYY/MM/DD HH:MM
            "%Y/%m/%d",  # YYYY/MM/DD
            "%Y.%m.%d %H:%M:%S",  # YYYY.MM.DD HH:MM:SS
            "%Y.%m.%d %H:%M",  # YYYY.MM.DD HH:MM
            "%Y.%m.%d",  # YYYY.MM.DD
            "%d.%m.%Y %H:%M:%S",  # DD.MM.YYYY HH:MM:SS
            "%d.%m.%Y %H:%M",  # DD.MM.YYYY HH:MM
            "%d.%m.%Y",  # DD.MM.YYYY
            "%Y-%m-%d",  # ISO corto
            "%Y-%m-%d %H:%M:%S",  # 2025-05-31 00:00:00
            "%-m/%-d/%y",  # 5/9/25 (sin ceros a la izquierda, año corto)
            "%m/%d/%y",  # 05/09/25 (con ceros a la izquierda, año corto)
        ]

        for fmt in formatos:
            try:
                return datetime.strptime(fecha_str, fmt)
            except ValueError:
                continue  # Intenta el siguiente formato

        # Si ninguno funciona, lanza un error con detalle
        raise ValueError(f"Formato de fecha inválido: {fecha_str}")

    def parse_tipo_dato_report(
        self, type_value: str | list[str] | None, format_type: str | None
    ) -> DataTypeReport:
        if isinstance(type_value, list) and type_value:
            type_str = type_value[0]
        elif isinstance(type_value, str):
            type_str = type_value
        else:
            return DataTypeReport.TEXTO

        try:
            if format_type == "date-time" and type_str == DataTypeReport.TEXTO.value:
                return DataTypeReport.FECHA
            if type_str == DataTypeReport.MONEDA.value:
                return DataTypeReport.MONEDA
            if type_str == DataTypeReport.NUMERO.value:
                return DataTypeReport.NUMERO
            if type_str == DataTypeReport.FECHA.value:
                return DataTypeReport.FECHA
            if type_str == DataTypeReport.INTEGER.value:
                return DataTypeReport.INTEGER
            return DataTypeReport[type_str]
        except KeyError:
            return DataTypeReport.TEXTO

    def validate_data_and_erp(
        self, tipo_dato: TipoDato, origin_column: OrigenColumna
    ) -> bool:
        return tipo_dato == TipoDato.FECHA and origin_column in [
            OrigenColumna.ERP,
            OrigenColumna.SAT,
            OrigenColumna.CLEAN_DATA,
        ]

    def validate_data_and_clean_data(
        self, tipo_dato: TipoDato, origin_column: OrigenColumna
    ) -> bool:
        return tipo_dato != TipoDato.FECHA and origin_column in [
            OrigenColumna.ERP,
            OrigenColumna.CLEAN_DATA,
        ]

    def validate_moneda_and_sat(
        self, tipo_dato: TipoDato, origin_column: OrigenColumna
    ) -> bool:
        return tipo_dato == TipoDato.MONEDA and origin_column in [
            OrigenColumna.ERP,
            OrigenColumna.SAT,
            OrigenColumna.CLEAN_DATA,
        ]
